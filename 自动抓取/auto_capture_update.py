#! python
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import re, os


def find_all_text(file_list=[], path=os.getcwd()):
    for file in os.listdir(path):
        if file.endswith('.TXT'):
            file_list.append('.\\'+file)
    return file_list


def match_input_excel(files):
    for file in files:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='Number')
        ws.cell(row=1, column=2, value='Resolution')
        print('hello')
        with open(file, encoding='gb18030', mode='r', errors='ignore') as f:
            line = f.readlines()
            resolution_regex = re.compile(r': Resolution Check: Monitoring data, Resolution (\d{4,5})')
            result = resolution_regex.findall(str(line))
            num = len(result)
            row_num = 1
            for i in range(num):
                ws.cell(row=i + 2, column=2, value=int(result[i]))
                ws.cell(row=i + 2, column=1, value=i + 1)
                row_num += 1
            print(row_num)

            num = [int(x) for x in result]

            chart = LineChart()
            chart.title = 'Resolution Trending'
            chart.style = 18
            chart.shape = 40
            chart.x_axis.title = 'Number'
            chart.y_axis.title = 'Resolution'
            if len(num) >= 2:
                print(file)
                print(max(num),min(num))
                chart.y_axis.scaling.max = max(num) + 500
                chart.y_axis.scaling.min = min(num) - 300
            else:
                chart.y_axis.scaling.max = 11000
                chart.y_axis.scaling.min = 7000

            data = Reference(ws, min_row=1, max_row=row_num, min_col=2, max_col=2)
            chart.add_data(data, titles_from_data=True)

            style = chart.series[0]
            style.marker.symbol = "triangle"
            style.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
            style.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
            style.smooth = True

            ws.add_chart(chart, 'D2')
            wb.save(file.rstrip('.TXT') + '.xlsx')


if __name__ == '__main__':
    a = find_all_text()
    match_input_excel(a)
